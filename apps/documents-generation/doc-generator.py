import os
import io

from re import I
from tempfile import NamedTemporaryFile

from pathlib import Path
from regex import W
import streamlit as st
import xlsxwriter
import xlsxwriter.worksheet
import tempfile


from openpyxl import load_workbook
from openpyxl.drawing.image import Image


WORK_DIR = Path(__file__).parent
DATA_DIR = WORK_DIR / "templates"
TEMPLATE_PATH = DATA_DIR / "template-estimate.xlsx"
FILE_PATH = DATA_DIR / "last_estimate.xlsx"

template = load_workbook(filename=str(TEMPLATE_PATH))
template.save(FILE_PATH)
template.close()

workbook = load_workbook(filename=str(FILE_PATH))
worksheet = workbook.active
# worksheet = workbook.worksheets[0]


# workbook = xlsxwriter.Workbook(str(TEMPLATE_PATH))
# worksheet : xlsxwriter.worksheet.Worksheet = workbook.get_worksheet_by_name("Смета")
# worksheet : xlsxwriter.worksheet.Worksheet = workbook.worksheets()
print(worksheet)

if "show" not in st.session_state:
    st.session_state.show = set()
    st.session_state.buttons = True
    st.session_state.a_full_name = "Кузько Василий Александрович"
    st.session_state.b_full_name = "Иванов Иван Иванович"
    st.session_state.a_status = "Исполнительный директор 'ООО Зеленые руки'"
    st.session_state.a_inn = "7721581040"
    st.session_state.work_fields = [
        "DSN-040-009", "Очистка вручную поверхности потолков от набела", "м2",
        "1", "71,80", "71,80", "", "", "",
    ]
    st.session_state.prices = "1.138,52", "525,21"

st.title("Генерация сметы по шаблону")

st.image(str(DATA_DIR / "estimate-image.png"))

st.header("Проверьте персональные данные субъектов документа")

st.subheader("Заказчика")
st.text_input("ФИО", value=st.session_state.a_full_name)
st.text_input("ИНН", value=st.session_state.a_inn)
st.text_input("Должность", value=st.session_state.a_status)

st.subheader("Подрядчика")
st.text_input("ФИО", value=st.session_state.b_full_name)

st.subheader("Работы")
st.text_input("№1", value=st.session_state.work_fields)

# st.checkbox("Я согласен в обработке персональных данных в соответствии с ФЗ-152")
if st.button("Подтвердить, подписать документ"):
    worksheet["A1"] = f"Заказчик: {st.session_state.a_full_name}"
    worksheet["A2"] = f"Подрядчик: {st.session_state.b_full_name}"
    worksheet["H5"] = st.session_state.a_full_name
    worksheet["H6"] = st.session_state.a_inn
    worksheet["H8"] = st.session_state.a_status

    for cln_iter, row_data in enumerate(st.session_state.work_fields):
        worksheet.cell(25, 2 + cln_iter, row_data)

    worksheet["H38"] = st.session_state.prices[0]
    worksheet["I38"] = st.session_state.prices[1]

    img = Image(str(DATA_DIR / "esign_a.png"))
    img.width = 100
    img.height = 100
    worksheet.add_image(img, "B59")
    st.header("Данные заполнены, подпись поставлена")
    # workbook.worksheets[0] = worksheet

if st.button("Отправить"):
    st.write("Документ сохранен, отправлен на подтверждение")
    with tempfile.NamedTemporaryFile(delete=False) as f:
        workbook.save(DATA_DIR / "approved.xlsx")
        # workbook.close()

# streamlit run doc-generator.py
